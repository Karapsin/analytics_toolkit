from __future__ import annotations

from pathlib import Path
from typing import Literal

import pandas as pd
from openpyxl import load_workbook

Axis = Literal["dates", "labels", "segment"]


def break_into_tables(
    df: pd.DataFrame,
    label: str,
    value: str,
    output: str | Path,
    sheet: str,
    date: str | None = None,
    index: str | None = None,
    segment: str | None = None,
    rows: Axis = "labels",
    columns: Axis | None = None,
    break_by: Axis | None = None,
) -> list[pd.DataFrame]:
    """Format a long metrics dataframe as one or more Excel tables.

    Returns the tables that were written, which is useful for tests or callers
    that need the formatted dataframe objects as well as the Excel output.
    """
    columns_map = _columns_map(label=label, date=date, segment=segment)
    _validate_input(
        df=df,
        columns_map=columns_map,
        value=value,
        index=index,
        rows=rows,
        columns=columns,
        break_by=break_by,
    )

    source = _sort_source(df=df, columns_map=columns_map, index=index)
    tables = _build_tables(
        df=source,
        columns_map=columns_map,
        value=value,
        rows=rows,
        columns=columns,
        break_by=break_by,
    )

    _write_tables(tables=tables, output=Path(output), sheet=sheet, break_by=break_by)
    return [table for _, table in tables]


def _columns_map(
    label: str,
    date: str | None,
    segment: str | None,
) -> dict[Axis, str | None]:
    return {"labels": label, "dates": date, "segment": segment}


def _validate_input(
    df: pd.DataFrame,
    columns_map: dict[Axis, str | None],
    value: str,
    index: str | None,
    rows: Axis,
    columns: Axis | None,
    break_by: Axis | None,
) -> None:
    required_axes: list[Axis] = [rows]
    if columns is not None:
        required_axes.append(columns)
    if break_by is not None:
        required_axes.append(break_by)

    for axis in required_axes:
        if axis not in columns_map:
            raise ValueError("rows, columns, and break_by must be one of: 'dates', 'labels', 'segment'.")
        if columns_map[axis] is None:
            raise ValueError(f"{axis!r} was requested, but its column argument was not provided.")

    required_columns = {value}
    required_columns.update(column for column in columns_map.values() if column is not None)
    if index is not None:
        required_columns.add(index)

    missing = sorted(column for column in required_columns if column not in df.columns)
    if missing:
        raise ValueError(f"Input dataframe is missing required columns: {missing}.")

    if columns is not None and rows == columns:
        raise ValueError("rows and columns must refer to different axes.")
    if break_by is not None and break_by in {rows, columns}:
        raise ValueError("break_by must be different from rows and columns.")


def _sort_source(
    df: pd.DataFrame,
    columns_map: dict[Axis, str | None],
    index: str | None,
) -> pd.DataFrame:
    if index is None:
        return df

    label_column = columns_map["labels"]
    sort_columns = [index]
    if label_column is not None and label_column != index:
        sort_columns.append(label_column)
    return df.sort_values(sort_columns, kind="stable")


def _build_tables(
    df: pd.DataFrame,
    columns_map: dict[Axis, str | None],
    value: str,
    rows: Axis,
    columns: Axis | None,
    break_by: Axis | None,
) -> list[tuple[object | None, pd.DataFrame]]:
    break_column = columns_map[break_by] if break_by is not None else None

    if break_column is None:
        return [(None, _build_table(df=df, columns_map=columns_map, value=value, rows=rows, columns=columns))]

    tables: list[tuple[object | None, pd.DataFrame]] = []
    for break_value, table_df in df.groupby(break_column, sort=False, dropna=False):
        table = _build_table(df=table_df, columns_map=columns_map, value=value, rows=rows, columns=columns)
        tables.append((break_value, table))
    return tables


def _build_table(
    df: pd.DataFrame,
    columns_map: dict[Axis, str | None],
    value: str,
    rows: Axis,
    columns: Axis | None,
) -> pd.DataFrame:
    row_column = columns_map[rows]
    if row_column is None:
        raise ValueError(f"{rows!r} was requested, but its column argument was not provided.")

    if columns is None:
        duplicate_mask = df.duplicated(subset=[row_column], keep=False)
        if duplicate_mask.any():
            duplicates = df.loc[duplicate_mask, row_column].drop_duplicates().tolist()
            raise ValueError(f"Values are not unique for rows={rows!r}: {duplicates}.")
        return df[[row_column, value]].reset_index(drop=True)

    column_column = columns_map[columns]
    if column_column is None:
        raise ValueError(f"{columns!r} was requested, but its column argument was not provided.")

    duplicate_mask = df.duplicated(subset=[row_column, column_column], keep=False)
    if duplicate_mask.any():
        duplicates = (
            df.loc[duplicate_mask, [row_column, column_column]]
            .drop_duplicates()
            .to_dict(orient="records")
        )
        raise ValueError(
            f"Values are not unique for rows={rows!r} and columns={columns!r}: {duplicates}."
        )

    table = df.pivot(index=row_column, columns=column_column, values=value)
    return table.reset_index().rename_axis(columns=None)


def _write_tables(
    tables: list[tuple[object | None, pd.DataFrame]],
    output: Path,
    sheet: str,
    break_by: Axis | None,
) -> None:
    output.parent.mkdir(parents=True, exist_ok=True)
    if output.exists():
        workbook = load_workbook(output, read_only=True)
        try:
            if sheet in workbook.sheetnames:
                raise ValueError(f"Sheet {sheet!r} already exists in {str(output)!r}.")
        finally:
            workbook.close()
        mode = "a"
    else:
        mode = "w"

    writer_kwargs = {"engine": "openpyxl", "mode": mode}
    if mode == "a":
        writer_kwargs["if_sheet_exists"] = "overlay"

    with pd.ExcelWriter(output, **writer_kwargs) as writer:
        startrow = 0
        for break_value, table in tables:
            if break_by is not None:
                title = pd.DataFrame({f"{break_by}": [break_value]})
                title.to_excel(writer, sheet_name=sheet, index=False, header=False, startrow=startrow)
                startrow += 1

            table.to_excel(writer, sheet_name=sheet, index=False, startrow=startrow)
            startrow += len(table.index) + 3

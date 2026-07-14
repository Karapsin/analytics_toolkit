from __future__ import annotations

from decimal import Decimal
from pathlib import Path

import pandas as pd
import pytest
from openpyxl import load_workbook

from analytics_toolkit.excel import break_table, pivot_and_break_table


@pytest.mark.parametrize("invalid_df", [None, "table", 42, [], [pd.DataFrame(), "table"]])
def test_table_helpers_reject_invalid_dataframe_specs(
    invalid_df: object,
    tmp_path: Path,
) -> None:
    expected_error = ValueError if invalid_df == [] else TypeError

    with pytest.raises(expected_error, match="dataframe"):
        break_table(invalid_df, tmp_path / "invalid.xlsx")  # type: ignore[arg-type]


@pytest.mark.parametrize(
    ("kwargs", "message"),
    [
        ({"value": "missing"}, "missing required columns"),
        ({"value": "value", "columns": "metric"}, "different dataframe columns"),
        ({"value": []}, "at least one"),
        ({"value": ["value", 1]}, "sequence of column names"),
        ({"value": ["value", "value"]}, "duplicate columns"),
    ],
)
def test_pivot_and_break_table_rejects_invalid_column_specs(
    kwargs: dict[str, object],
    message: str,
    tmp_path: Path,
) -> None:
    df = pd.DataFrame({"metric": ["users"], "value": [1]})

    with pytest.raises(ValueError, match=message):
        pivot_and_break_table(
            df,
            rows="metric",
            output=tmp_path / "invalid-columns.xlsx",
            **kwargs,  # type: ignore[arg-type]
        )


def test_pivot_and_break_table_rejects_multi_value_row_role_collisions(tmp_path: Path) -> None:
    df = pd.DataFrame({"group": ["A"], "metric": ["existing"], "users": [1], "orders": [2]})

    with pytest.raises(ValueError, match="grouping column"):
        pivot_and_break_table(
            df,
            rows="group",
            value=["users", "orders"],
            columns="group",
            output=tmp_path / "group-collision.xlsx",
        )
    with pytest.raises(ValueError, match="already exists"):
        pivot_and_break_table(
            df,
            rows="metric",
            value=["users", "orders"],
            output=tmp_path / "existing-row.xlsx",
        )


def test_pivot_without_columns_rejects_conflicting_rows(tmp_path: Path) -> None:
    df = pd.DataFrame({"metric": ["users", "users"], "value": [1, 2]})

    with pytest.raises(ValueError, match="not unique"):
        pivot_and_break_table(
            df,
            rows="metric",
            value="value",
            output=tmp_path / "conflicting-rows.xlsx",
        )


def test_pivot_without_columns_preserves_row_order(tmp_path: Path) -> None:
    df = pd.DataFrame({"metric": ["orders", "users"], "value": [2, 1]})

    tables = pivot_and_break_table(
        df,
        rows="metric",
        value="value",
        output=tmp_path / "no-columns.xlsx",
    )

    assert tables[None][0].to_dict(orient="records") == [
        {"metric": "orders", "value": 2},
        {"metric": "users", "value": 1},
    ]


def test_break_table_rejects_missing_and_colliding_group_columns(tmp_path: Path) -> None:
    df = pd.DataFrame({"metric": ["users"]})

    with pytest.raises(ValueError, match="missing required columns"):
        break_table(df, tmp_path / "missing-break.xlsx", break_by="missing")
    with pytest.raises(ValueError, match="different dataframe columns"):
        break_table(df, tmp_path / "colliding-break.xlsx", break_by="metric", sheet_by="metric")


def test_pivot_and_break_table_writes_multiple_sheets_and_tables(tmp_path: Path) -> None:
    df = pd.DataFrame(
        {
            "metric": ["users", "users", "arpu", "arpu", "users", "users", "arpu", "arpu"],
            "ab_group": ["control", "test_1", "control", "test_1", "control", "test_1", "control", "test_1"],
            "qr_group": ["ALL", "ALL", "ALL", "ALL", "1", "1", "1", "1"],
            "start_dt": ["2026-03-30", "2026-03-30", "2026-03-30", "2026-03-30", "2026-04-01", "2026-04-01", "2026-04-01", "2026-04-01"],
            "value": [100, 110, 2.5, 2.7, 50, 55, 1.1, 1.2],
        }
    )

    output = tmp_path / "report.xlsx"
    tables = pivot_and_break_table(
        df=df,
        rows="metric",
        value="value",
        output=output,
        columns="ab_group",
        break_by="qr_group",
        sheet_by="start_dt",
    )

    assert list(tables) == ["2026-03-30", "2026-04-01"]
    assert len(tables["2026-03-30"]) == 1
    assert tables["2026-03-30"][0].columns.tolist() == ["metric", "control", "test_1"]

    workbook = load_workbook(output, read_only=True, data_only=True)
    try:
        assert workbook.sheetnames == ["2026-03-30", "2026-04-01"]
        first_sheet = workbook["2026-03-30"]
        rows = list(first_sheet.iter_rows(values_only=True))
        assert rows[:4] == [
            ("ALL", None, None),
            ("metric", "control", "test_1"),
            ("users", 100, 110),
            ("arpu", 2.5, 2.7),
        ]
    finally:
        workbook.close()


def test_pivot_and_break_table_sanitizes_and_deduplicates_sheet_names(tmp_path: Path) -> None:
    df = pd.DataFrame(
        {
            "metric": ["users", "users", "users", "users"],
            "ab_group": ["control", "test_1", "control", "test_1"],
            "sheet_bucket": [
                "Report/Name:One*?",
                "Report/Name:One*?",
                "X" * 40,
                "X" * 40 + " trailing",
            ],
            "value": [1, 2, 3, 4],
        }
    )

    output = tmp_path / "sanitized.xlsx"
    pivot_and_break_table(
        df=df,
        rows="metric",
        value="value",
        output=output,
        columns="ab_group",
        sheet_by="sheet_bucket",
    )

    workbook = load_workbook(output, read_only=True, data_only=True)
    try:
        assert workbook.sheetnames == [
            "Report_Name_One__",
            "XXXXXXXXXXXXXXXXXXXXXXXXXXXXXXX",
            "XXXXXXXXXXXXXXXXXXXXXXXXXXX (2)",
        ]
    finally:
        workbook.close()


def test_pivot_and_break_table_rejects_duplicates_within_group_slices(tmp_path: Path) -> None:
    df = pd.DataFrame(
        {
            "metric": ["users", "users"],
            "ab_group": ["control", "control"],
            "start_dt": ["2026-03-30", "2026-03-30"],
            "value": [100, 120],
        }
    )

    with pytest.raises(ValueError, match="not unique"):
        pivot_and_break_table(
            df=df,
            rows="metric",
            value="value",
            output=tmp_path / "duplicates.xlsx",
            columns="ab_group",
            sheet_by="start_dt",
        )


def test_pivot_and_break_table_accepts_multiple_value_columns(tmp_path: Path) -> None:
    df = pd.DataFrame(
        {
            "start_dt": ["2026-03-30", "2026-03-30"],
            "qr_group": ["ALL", "ALL"],
            "ab_group": ["control", "test_1"],
            "users": [100, 110],
            "arpu": [2.5, 2.7],
        }
    )

    output = tmp_path / "multi_value.xlsx"
    tables = pivot_and_break_table(
        df=df,
        rows="metric",
        value=["users", "arpu"],
        output=output,
        columns="ab_group",
        break_by="qr_group",
        sheet_by="start_dt",
    )

    assert tables["2026-03-30"][0].to_dict(orient="records") == [
        {"metric": "users", "control": 100.0, "test_1": 110.0},
        {"metric": "arpu", "control": 2.5, "test_1": 2.7},
    ]

    workbook = load_workbook(output, read_only=True, data_only=True)
    try:
        rows = list(workbook["2026-03-30"].iter_rows(values_only=True))
        assert rows[:4] == [
            ("ALL", None, None),
            ("metric", "control", "test_1"),
            ("users", 100, 110),
            ("arpu", 2.5, 2.7),
        ]
    finally:
        workbook.close()


def test_pivot_and_break_table_detects_value_columns_when_omitted(tmp_path: Path) -> None:
    df = pd.DataFrame(
        {
            "start_dt": ["2026-03-30", "2026-03-30"],
            "qr_group": ["ALL", "ALL"],
            "ab_group": ["control", "test_1"],
            "users": [100, 110],
            "arpu": [2.5, 2.7],
        }
    )

    output = tmp_path / "auto_value.xlsx"
    tables = pivot_and_break_table(
        df=df,
        rows="metric",
        output=output,
        columns="ab_group",
        break_by="qr_group",
        sheet_by="start_dt",
    )

    assert tables["2026-03-30"][0].to_dict(orient="records") == [
        {"metric": "users", "control": 100.0, "test_1": 110.0},
        {"metric": "arpu", "control": 2.5, "test_1": 2.7},
    ]


def test_pivot_and_break_table_accepts_multiple_dataframes_side_by_side(tmp_path: Path) -> None:
    first_df = pd.DataFrame(
        {
            "metric": ["users", "users"],
            "ab_group": ["control", "test_1"],
            "qr_group": ["ALL", "ALL"],
            "start_dt": ["2026-03-30", "2026-03-30"],
            "value": [100, 110],
        }
    )
    second_df = pd.DataFrame(
        {
            "metric": ["arpu", "arpu"],
            "ab_group": ["control", "test_1"],
            "qr_group": ["ALL", "ALL"],
            "start_dt": ["2026-03-30", "2026-03-30"],
            "value": [2.5, 2.7],
        }
    )

    output = tmp_path / "multi_df_pivot.xlsx"
    tables = pivot_and_break_table(
        df=[first_df, second_df],
        rows="metric",
        value="value",
        output=output,
        columns="ab_group",
        break_by="qr_group",
        sheet_by="start_dt",
    )

    assert tables["2026-03-30"][0][0].to_dict(orient="records") == [
        {"metric": "users", "control": 100, "test_1": 110},
    ]
    assert tables["2026-03-30"][1][0].to_dict(orient="records") == [
        {"metric": "arpu", "control": 2.5, "test_1": 2.7},
    ]

    workbook = load_workbook(output, read_only=True, data_only=True)
    try:
        rows = list(workbook["2026-03-30"].iter_rows(values_only=True))
        assert rows[:3] == [
            ("ALL", None, None, None, "ALL", None, None),
            ("metric", "control", "test_1", None, "metric", "control", "test_1"),
            ("users", 100, 110, None, "arpu", 2.5, 2.7),
        ]
    finally:
        workbook.close()


def test_pivot_and_break_table_enforces_same_row_order_with_padding(tmp_path: Path) -> None:
    first_df = pd.DataFrame(
        {
            "metric": ["users", "users", "arpu", "arpu"],
            "ab_group": ["control", "test_1", "control", "test_1"],
            "start_dt": ["2026-03-30"] * 4,
            "value": [100, 110, 2.5, 2.7],
        }
    )
    second_df = pd.DataFrame(
        {
            "metric": ["arpu", "arpu"],
            "ab_group": ["control", "test_1"],
            "start_dt": ["2026-03-30", "2026-03-30"],
            "value": [3.1, 3.3],
        }
    )

    output = tmp_path / "enforced_row_order_padding.xlsx"
    tables = pivot_and_break_table(
        df=[first_df, second_df],
        rows="metric",
        value="value",
        output=output,
        columns="ab_group",
        sheet_by="start_dt",
        enforce_same_row_order=True,
    )

    assert tables["2026-03-30"][0][0].to_dict(orient="records") == [
        {"metric": "users", "control": 100.0, "test_1": 110.0},
        {"metric": "arpu", "control": 2.5, "test_1": 2.7},
    ]
    second_table = tables["2026-03-30"][1][0]
    assert second_table["metric"].tolist() == ["users", "arpu"]
    assert pd.isna(second_table.iloc[0]["control"])
    assert pd.isna(second_table.iloc[0]["test_1"])
    assert second_table.iloc[1]["control"] == 3.1
    assert second_table.iloc[1]["test_1"] == 3.3

    workbook = load_workbook(output, read_only=True, data_only=True)
    try:
        rows = list(workbook["2026-03-30"].iter_rows(values_only=True))
        assert rows[:3] == [
            ("metric", "control", "test_1", None, "metric", "control", "test_1"),
            ("users", 100, 110, None, "users", None, None),
            ("arpu", 2.5, 2.7, None, "arpu", 3.1, 3.3),
        ]
    finally:
        workbook.close()


def test_pivot_and_break_table_enforces_same_row_order_by_reordering(tmp_path: Path) -> None:
    first_df = pd.DataFrame(
        {
            "metric": ["users", "users", "arpu", "arpu"],
            "ab_group": ["control", "test_1", "control", "test_1"],
            "start_dt": ["2026-03-30"] * 4,
            "value": [100, 110, 2.5, 2.7],
        }
    )
    second_df = pd.DataFrame(
        {
            "metric": ["arpu", "arpu", "users", "users"],
            "ab_group": ["control", "test_1", "control", "test_1"],
            "start_dt": ["2026-03-30"] * 4,
            "value": [3.1, 3.3, 200, 220],
        }
    )

    output = tmp_path / "enforced_row_order_reordered.xlsx"
    tables = pivot_and_break_table(
        df=[first_df, second_df],
        rows="metric",
        value="value",
        output=output,
        columns="ab_group",
        sheet_by="start_dt",
        enforce_same_row_order=True,
    )

    assert tables["2026-03-30"][1][0].to_dict(orient="records") == [
        {"metric": "users", "control": 200.0, "test_1": 220.0},
        {"metric": "arpu", "control": 3.1, "test_1": 3.3},
    ]


def test_pivot_and_break_table_rejects_extra_rows_when_enforcing_same_row_order(
    tmp_path: Path,
) -> None:
    first_df = pd.DataFrame(
        {
            "metric": ["users", "users"],
            "ab_group": ["control", "test_1"],
            "start_dt": ["2026-03-30", "2026-03-30"],
            "value": [100, 110],
        }
    )
    second_df = pd.DataFrame(
        {
            "metric": ["users", "users", "arpu", "arpu"],
            "ab_group": ["control", "test_1", "control", "test_1"],
            "start_dt": ["2026-03-30"] * 4,
            "value": [200, 220, 3.1, 3.3],
        }
    )

    with pytest.raises(ValueError, match="extra row labels"):
        pivot_and_break_table(
            df=[first_df, second_df],
            rows="metric",
            value="value",
            output=tmp_path / "enforced_row_order_error.xlsx",
            columns="ab_group",
            sheet_by="start_dt",
            enforce_same_row_order=True,
        )


def test_break_table_writes_grouped_raw_tables_without_uniqueness_checks(tmp_path: Path) -> None:
    df = pd.DataFrame(
        {
            "metric": ["users", "users", "users", "users"],
            "ab_group": ["control", "control", "test_1", "test_1"],
            "qr_group": ["ALL", "ALL", "ALL", "ALL"],
            "start_dt": ["2026-03-30"] * 4,
            "value": [100, 120, 110, 130],
        }
    )

    output = tmp_path / "raw_tables.xlsx"
    tables = break_table(
        df=df,
        output=output,
        break_by="qr_group",
        sheet_by="start_dt",
    )

    assert list(tables) == ["2026-03-30"]
    assert tables["2026-03-30"][0].to_dict(orient="records") == [
        {"metric": "users", "ab_group": "control", "value": 100},
        {"metric": "users", "ab_group": "control", "value": 120},
        {"metric": "users", "ab_group": "test_1", "value": 110},
        {"metric": "users", "ab_group": "test_1", "value": 130},
    ]

    workbook = load_workbook(output, read_only=True, data_only=True)
    try:
        assert workbook.sheetnames == ["2026-03-30"]
        rows = list(workbook["2026-03-30"].iter_rows(values_only=True))
        assert rows[:6] == [
            ("ALL", None, None),
            ("metric", "ab_group", "value"),
            ("users", "control", 100),
            ("users", "control", 120),
            ("users", "test_1", 110),
            ("users", "test_1", 130),
        ]
    finally:
        workbook.close()


def test_break_table_accepts_multiple_dataframes_side_by_side(tmp_path: Path) -> None:
    first_df = pd.DataFrame(
        {
            "metric": ["users", "users"],
            "ab_group": ["control", "test_1"],
            "qr_group": ["ALL", "ALL"],
            "start_dt": ["2026-03-30", "2026-03-30"],
            "value": [100, 110],
        }
    )
    second_df = pd.DataFrame(
        {
            "metric": ["arpu", "arpu"],
            "ab_group": ["control", "test_1"],
            "qr_group": ["ALL", "ALL"],
            "start_dt": ["2026-03-30", "2026-03-30"],
            "value": [2.5, 2.7],
        }
    )

    output = tmp_path / "multi_df_raw.xlsx"
    tables = break_table(
        df=[first_df, second_df],
        output=output,
        break_by="qr_group",
        sheet_by="start_dt",
    )

    assert tables["2026-03-30"][0][0].to_dict(orient="records") == [
        {"metric": "users", "ab_group": "control", "value": 100},
        {"metric": "users", "ab_group": "test_1", "value": 110},
    ]
    assert tables["2026-03-30"][1][0].to_dict(orient="records") == [
        {"metric": "arpu", "ab_group": "control", "value": 2.5},
        {"metric": "arpu", "ab_group": "test_1", "value": 2.7},
    ]

    workbook = load_workbook(output, read_only=True, data_only=True)
    try:
        rows = list(workbook["2026-03-30"].iter_rows(values_only=True))
        assert rows[:4] == [
            ("ALL", None, None, None, "ALL", None, None),
            ("metric", "ab_group", "value", None, "metric", "ab_group", "value"),
            ("users", "control", 100, None, "arpu", "control", 2.5),
            ("users", "test_1", 110, None, "arpu", "test_1", 2.7),
        ]
    finally:
        workbook.close()


def test_break_table_aligns_headers_by_table_index_across_multiple_dataframes(tmp_path: Path) -> None:
    first_df = pd.DataFrame(
        {
            "metric": ["users", "orders", "users"],
            "ab_group": ["control", "control", "test_1"],
            "qr_group": ["ALL", "ALL", "B"],
            "start_dt": ["2026-03-30", "2026-03-30", "2026-03-30"],
            "value": [100, 50, 110],
        }
    )
    second_df = pd.DataFrame(
        {
            "metric": ["arpu", "margin", "arpu", "margin"],
            "ab_group": ["control", "control", "test_1", "test_1"],
            "qr_group": ["ALL", "ALL", "B", "B"],
            "start_dt": ["2026-03-30"] * 4,
            "value": [2.5, 0.2, 2.7, 0.25],
        }
    )

    output = tmp_path / "aligned_multi_df_raw.xlsx"
    break_table(
        df=[first_df, second_df],
        output=output,
        break_by="qr_group",
        sheet_by="start_dt",
    )

    workbook = load_workbook(output, read_only=True, data_only=True)
    try:
        rows = list(workbook["2026-03-30"].iter_rows(values_only=True))
        assert rows[0] == ("ALL", None, None, None, "ALL", None, None)
        assert rows[1] == ("metric", "ab_group", "value", None, "metric", "ab_group", "value")
        assert rows[6] == ("B", None, None, None, "B", None, None)
        assert rows[7] == ("metric", "ab_group", "value", None, "metric", "ab_group", "value")
    finally:
        workbook.close()


def test_pivot_and_break_table_replaces_existing_workbook_by_default(tmp_path: Path) -> None:
    output = tmp_path / "replace.xlsx"

    first_df = pd.DataFrame(
        {
            "metric": ["users", "users"],
            "ab_group": ["control", "test_1"],
            "start_dt": ["2026-03-30", "2026-03-30"],
            "value": [100, 110],
        }
    )
    second_df = pd.DataFrame(
        {
            "metric": ["arpu", "arpu"],
            "ab_group": ["control", "test_1"],
            "start_dt": ["2026-04-01", "2026-04-01"],
            "value": [2.5, 2.7],
        }
    )

    pivot_and_break_table(
        df=first_df,
        rows="metric",
        value="value",
        output=output,
        columns="ab_group",
        sheet_by="start_dt",
    )
    pivot_and_break_table(
        df=second_df,
        rows="metric",
        value="value",
        output=output,
        columns="ab_group",
        sheet_by="start_dt",
    )

    workbook = load_workbook(output, read_only=True, data_only=True)
    try:
        assert workbook.sheetnames == ["2026-04-01"]
        rows = list(workbook["2026-04-01"].iter_rows(values_only=True))
        assert rows[:2] == [
            ("metric", "control", "test_1"),
            ("arpu", 2.5, 2.7),
        ]
    finally:
        workbook.close()


def test_pivot_and_break_table_appends_new_sheets_when_requested(tmp_path: Path) -> None:
    output = tmp_path / "append.xlsx"

    first_df = pd.DataFrame(
        {
            "metric": ["users", "users"],
            "ab_group": ["control", "test_1"],
            "start_dt": ["2026-03-30", "2026-03-30"],
            "value": [100, 110],
        }
    )
    second_df = pd.DataFrame(
        {
            "metric": ["arpu", "arpu"],
            "ab_group": ["control", "test_1"],
            "start_dt": ["2026-04-01", "2026-04-01"],
            "value": [2.5, 2.7],
        }
    )

    pivot_and_break_table(
        df=first_df,
        rows="metric",
        value="value",
        output=output,
        columns="ab_group",
        sheet_by="start_dt",
    )
    pivot_and_break_table(
        df=second_df,
        rows="metric",
        value="value",
        output=output,
        columns="ab_group",
        sheet_by="start_dt",
        append=True,
    )

    workbook = load_workbook(output, read_only=True, data_only=True)
    try:
        assert workbook.sheetnames == ["2026-03-30", "2026-04-01"]
    finally:
        workbook.close()


def test_break_table_writes_decimal_values_as_numeric_excel_cells(tmp_path: Path) -> None:
    df = pd.DataFrame(
        {
            "metric": ["arpu"],
            "ab_group": ["control"],
            "qr_group": ["ALL"],
            "start_dt": ["2026-03-30"],
            "value": [Decimal("72.4867078207333238")],
        }
    )

    output = tmp_path / "decimal_raw.xlsx"
    break_table(
        df=df,
        output=output,
        break_by="qr_group",
        sheet_by="start_dt",
    )

    workbook = load_workbook(output, read_only=False, data_only=True)
    try:
        cell = workbook["2026-03-30"]["C3"]
        assert cell.data_type == "n"
        assert cell.value == pytest.approx(72.4867078207333238)
    finally:
        workbook.close()


def test_break_table_prettify_formats_numeric_cells_by_body_row(tmp_path: Path) -> None:
    df = pd.DataFrame(
        {
            "metric": ["conversion", "mixed", "margin", "users", "comment"],
            "note": ["all_numeric", "has_text", "bounded", "large", "text_only"],
            "control": [0.125, "n/a", -10, 1000, "n/a"],
            "test_1": [0.75, 0.4, 100, 50, "missing"],
        }
    )

    output = tmp_path / "prettified_raw.xlsx"
    tables = break_table(
        df=df,
        output=output,
        prettify=True,
    )

    assert tables[None][0].equals(df)

    workbook = load_workbook(output, read_only=False, data_only=True)
    try:
        sheet = workbook["Sheet1"]
        assert sheet["A2"].number_format == "General"
        assert sheet["B2"].number_format == "General"
        assert sheet["C2"].number_format == "0.00%"
        assert sheet["D2"].number_format == "0.00%"

        assert sheet["C3"].number_format == "General"
        assert sheet["D3"].number_format == "0.00%"

        assert sheet["C4"].number_format == "0.00"
        assert sheet["D4"].number_format == "0.00"

        assert sheet["C5"].number_format == "#,##0"
        assert sheet["D5"].number_format == "#,##0"

        assert sheet["C6"].number_format == "General"
        assert sheet["D6"].number_format == "General"
    finally:
        workbook.close()


def test_break_table_does_not_apply_custom_formats_by_default(tmp_path: Path) -> None:
    df = pd.DataFrame(
        {
            "metric": ["conversion", "users"],
            "control": [0.125, 1000],
            "test_1": [0.75, 50],
        }
    )

    output = tmp_path / "plain_raw.xlsx"
    break_table(
        df=df,
        output=output,
    )

    workbook = load_workbook(output, read_only=False, data_only=True)
    try:
        sheet = workbook["Sheet1"]
        assert sheet["B2"].number_format == "General"
        assert sheet["C2"].number_format == "General"
        assert sheet["B3"].number_format == "General"
        assert sheet["C3"].number_format == "General"
    finally:
        workbook.close()


def test_pivot_and_break_table_prettify_formats_pivoted_body_rows(tmp_path: Path) -> None:
    df = pd.DataFrame(
        {
            "metric": ["conversion", "conversion", "margin", "margin", "users", "users"],
            "ab_group": ["control", "test_1", "control", "test_1", "control", "test_1"],
            "value": [0.125, 0.75, -10, 100, 1000, 50],
        }
    )

    output = tmp_path / "prettified_pivot.xlsx"
    tables = pivot_and_break_table(
        df=df,
        rows="metric",
        value="value",
        output=output,
        columns="ab_group",
        prettify=True,
    )

    assert tables[None][0].to_dict(orient="records") == [
        {"metric": "conversion", "control": 0.125, "test_1": 0.75},
        {"metric": "margin", "control": -10.0, "test_1": 100.0},
        {"metric": "users", "control": 1000.0, "test_1": 50.0},
    ]

    workbook = load_workbook(output, read_only=False, data_only=True)
    try:
        sheet = workbook["Sheet1"]
        assert sheet["A2"].number_format == "General"
        assert sheet["B2"].number_format == "0.00%"
        assert sheet["C2"].number_format == "0.00%"
        assert sheet["B3"].number_format == "0.00"
        assert sheet["C3"].number_format == "0.00"
        assert sheet["B4"].number_format == "#,##0"
        assert sheet["C4"].number_format == "#,##0"
    finally:
        workbook.close()


def test_pivot_and_break_table_writes_decimal_values_as_numeric_excel_cells(
    tmp_path: Path,
) -> None:
    df = pd.DataFrame(
        {
            "metric": ["arpu", "arpu"],
            "ab_group": ["control", "test_1"],
            "start_dt": ["2026-03-30", "2026-03-30"],
            "value": [Decimal("72.4867078207333238"), Decimal("70.6603563524410553")],
        }
    )

    output = tmp_path / "decimal_pivot.xlsx"
    pivot_and_break_table(
        df=df,
        rows="metric",
        value="value",
        output=output,
        columns="ab_group",
        sheet_by="start_dt",
    )

    workbook = load_workbook(output, read_only=False, data_only=True)
    try:
        control_cell = workbook["2026-03-30"]["B2"]
        test_cell = workbook["2026-03-30"]["C2"]
        assert control_cell.data_type == "n"
        assert test_cell.data_type == "n"
        assert control_cell.value == pytest.approx(72.4867078207333238)
        assert test_cell.value == pytest.approx(70.6603563524410553)
    finally:
        workbook.close()


def test_enforced_order_rejects_extra_and_misaligned_break_groups(tmp_path: Path) -> None:
    first = pd.DataFrame(
        {"metric": ["users"], "break": ["A"], "value": [1]}
    )
    extra = pd.DataFrame(
        {"metric": ["users", "users"], "break": ["A", "B"], "value": [2, 3]}
    )
    misaligned = pd.DataFrame(
        {"metric": ["users"], "break": ["B"], "value": [2]}
    )

    with pytest.raises(ValueError, match="more tables"):
        pivot_and_break_table(
            [first, extra],
            rows="metric",
            value="value",
            output=tmp_path / "extra-groups.xlsx",
            break_by="break",
            enforce_same_row_order=True,
        )
    with pytest.raises(ValueError, match="do not align"):
        pivot_and_break_table(
            [first, misaligned],
            rows="metric",
            value="value",
            output=tmp_path / "misaligned-groups.xlsx",
            break_by="break",
            enforce_same_row_order=True,
        )


def test_break_table_places_sparse_dataframe_groups_at_stable_coordinates(tmp_path: Path) -> None:
    first = pd.DataFrame(
        {"metric": ["users", "orders"], "break": ["A", "B"], "value": [1, 2]}
    )
    second = pd.DataFrame({"metric": ["margin"], "break": ["A"], "value": [0.5]})

    output = tmp_path / "sparse-groups.xlsx"
    break_table([first, second], output, break_by="break")

    workbook = load_workbook(output, read_only=False, data_only=True)
    try:
        sheet = workbook["Sheet1"]
        assert sheet["A1"].value == "A"
        assert sheet["D1"].value == "A"
        assert sheet["A6"].value == "B"
        assert sheet["D6"].value is None
    finally:
        workbook.close()


def test_break_table_handles_disjoint_sheets_and_special_numeric_values(tmp_path: Path) -> None:
    first = pd.DataFrame(
        {"sheet": ["first"], "label": ["flag"], "flag": [True], "missing": [float("nan")]}
    )
    second = pd.DataFrame(
        {"sheet": ["second"], "label": ["ratio"], "flag": [0.5], "missing": [None]}
    )

    output = tmp_path / "disjoint-sheets.xlsx"
    break_table([first, second], output, sheet_by="sheet", prettify=True)

    workbook = load_workbook(output, read_only=False, data_only=True)
    try:
        assert workbook.sheetnames == ["first", "second"]
        assert workbook["first"]["B2"].number_format == "General"
        assert workbook["first"]["C2"].number_format == "General"
        assert workbook["second"]["B2"].number_format == "0.00%"
    finally:
        workbook.close()


def test_sheet_names_cover_missing_empty_and_multiple_deduplications(tmp_path: Path) -> None:
    df = pd.DataFrame(
        {
            "sheet": [None, "''", "X" * 40, "X" * 39 + "A", "X" * 39 + "B"],
            "value": [1, 2, 3, 4, 5],
        }
    )

    output = tmp_path / "edge-sheet-names.xlsx"
    break_table(df, output, sheet_by="sheet")

    workbook = load_workbook(output, read_only=True, data_only=True)
    try:
        expected_names = [
            "sheet_NA",
            "sheet_value",
            "XXXXXXXXXXXXXXXXXXXXXXXXXXXXXXX",
            "XXXXXXXXXXXXXXXXXXXXXXXXXXX (2)",
            "XXXXXXXXXXXXXXXXXXXXXXXXXXX (3)",
        ]
        assert set(workbook.sheetnames) == set(expected_names)
        assert [name for name in workbook.sheetnames if name.startswith("X")] == expected_names[2:]
    finally:
        workbook.close()

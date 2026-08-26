from __future__ import annotations

from types import SimpleNamespace
from typing import TYPE_CHECKING, Any

import analytics_toolkit.ab_utils as ab_utils_module
import pandas as pd
import pytest
from analytics_toolkit.ab_utils import metrics as metrics_module
from analytics_toolkit.ab_utils import reporting
from openpyxl import load_workbook

if TYPE_CHECKING:
    from pathlib import Path


def _table_info(table: str) -> SimpleNamespace:
    return SimpleNamespace(
        exists=True,
        backend="gp",
        table=table,
        resolved_table=table,
        columns={
            "user_id": "bigint",
            "group_name": "text",
            "segment_name": "text",
            "orders": "bigint",
            "revenue": "numeric",
            "views": "bigint",
        },
    )


def _metric_row(segment: object, test_group: str) -> pd.DataFrame:
    return pd.DataFrame(
        {
            "segment_name": [segment],
            "metric_type": ["mean"],
            "group_1": [test_group],
            "group_2": ["control"],
            "metric_name": ["orders"],
            "n_group_2": [100],
            "n_group_1": [90],
            "outliers_cutoff": [10.0],
            "outliers_n_group_2": [0],
            "outliers_n_group_1": [0],
            "metric_group_2": [2.0],
            "metric_group_1": [2.2],
            "variance_group_2": [1.0],
            "variance_group_1": [1.1],
            "delta_abs": [0.2],
            "delta_relative": [0.1],
            "mde_abs": [0.15],
            "mde_relative": [0.075],
            "s.e.": [0.05],
            "p-value": [0.005],
        }
    )


def _install_sql_fakes(
    monkeypatch: pytest.MonkeyPatch,
    *,
    captured: dict[str, Any],
) -> None:
    monkeypatch.setattr(reporting.sql, "table_info", lambda _db_key, table: _table_info(table))

    def fake_read(_db_key: str, query: str, **_kwargs: object) -> pd.DataFrame:
        captured.setdefault("read_queries", []).append(query)
        if reporting._SEGMENT_VALUE_COLUMN in query:
            return pd.DataFrame({reporting._SEGMENT_VALUE_COLUMN: ["north"]})
        return pd.DataFrame({"__analytics_toolkit_group_value__": ["control", "test_1"]})

    monkeypatch.setattr(reporting.sql, "read", fake_read)

    def fake_compute(
        db_key: str,
        tasks: dict[str, dict[str, object]],
        **kwargs: object,
    ) -> dict[str, pd.DataFrame]:
        captured.update({"db_key": db_key, "tasks": tasks, "kwargs": kwargs})
        frames: dict[str, pd.DataFrame] = {}
        for name, spec in tasks.items():
            labels = spec.get("labels", {})
            frame = _metric_row(
                labels.get("segment_name") if isinstance(labels, dict) else None,
                "test_all" if name.endswith("pooled") else "test_1",
            )
            if not labels:
                frame = frame.drop(columns="segment_name")
            frames[name] = frame
        return frames

    monkeypatch.setattr(reporting, "compute_test_metrics_sql_native", fake_compute)


def test_compute_metrics_report_is_exported() -> None:
    assert ab_utils_module.compute_metrics_report is reporting.compute_metrics_report
    assert metrics_module.compute_metrics_report is reporting.compute_metrics_report
    assert not hasattr(ab_utils_module, "compute_segment_metrics_report")
    assert not hasattr(metrics_module, "compute_segment_metrics_report")
    assert not hasattr(reporting, "compute_segment_metrics_report")


def test_resolve_group_order_appends_omitted_observed_groups() -> None:
    assert reporting._resolve_group_order(
        groups_order=["control", "test_all"],
        observed_groups=["control", "test_3", "test_2", "test_all"],
    ) == ["control", "test_all", "test_3", "test_2"]


def test_compute_metrics_report_builds_segment_and_pooled_tasks(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    captured: dict[str, Any] = {}
    _install_sql_fakes(monkeypatch, captured=captured)
    monkeypatch.setattr(
        reporting.excel,
        "break_table",
        lambda *_args, **_kwargs: pytest.fail("Excel must be disabled"),
    )

    result = reporting.compute_metrics_report(
        "sandbox.user_metrics",
        "segment_name",
        db_key="analytics",
        sql_where="experiment_id = 7",
        ratio_metrics=[
            {"name": "revenue_per_view", "numerator": "revenue", "denominator": "views"}
        ],
        groups_order=["control", "test_1", "test_all"],
        create_excel=False,
        outliers_quantile=0.95,
        concurrency=2,
    )

    assert result["segment_name"].tolist() == ["ALL", "ALL", "north", "north"]
    assert result["group_1"].tolist() == ["test_1", "test_all", "test_1", "test_all"]
    assert captured["db_key"] == "analytics"
    assert captured["kwargs"]["metric_columns"] == ["orders", "revenue", "views"]
    assert captured["kwargs"]["outliers_quantile"] == 0.95
    assert captured["kwargs"]["concurrency"] == 2

    tasks = captured["tasks"]
    assert list(tasks) == [
        "segment_0000_groups",
        "segment_0000_pooled",
        "segment_0001_groups",
        "segment_0001_pooled",
    ]
    assert tasks["segment_0000_groups"]["sql_where"] == "(experiment_id = 7)"
    assert "segment_name" in str(tasks["segment_0001_groups"]["sql_where"])
    assert "north" in str(tasks["segment_0001_groups"]["sql_where"])
    assert "CASE WHEN" in str(tasks["segment_0000_pooled"]["source"])
    assert "test_all" in str(tasks["segment_0000_pooled"]["source"])


def test_compute_metrics_report_without_segment_builds_only_total_tasks(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    captured: dict[str, Any] = {}
    _install_sql_fakes(monkeypatch, captured=captured)

    result = reporting.compute_metrics_report(
        "sandbox.user_metrics",
        db_key="analytics",
        sql_where="experiment_id = 7",
        create_excel=False,
    )

    assert "segment_name" not in result.columns
    assert result["group_1"].tolist() == ["test_1", "test_all"]
    assert list(captured["tasks"]) == ["total_groups", "total_pooled"]
    assert captured["tasks"]["total_groups"]["sql_where"] == "(experiment_id = 7)"
    assert "labels" not in captured["tasks"]["total_groups"]
    assert "segment_name" not in str(captured["tasks"]["total_pooled"]["source"])
    assert all(reporting._SEGMENT_VALUE_COLUMN not in query for query in captured["read_queries"])
    assert captured["kwargs"]["metric_columns"] == ["orders", "revenue", "views"]


def test_compute_metrics_report_writes_ordered_named_workbook(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    captured: dict[str, Any] = {}
    _install_sql_fakes(monkeypatch, captured=captured)
    output = tmp_path / "custom.xlsx"

    result = reporting.compute_metrics_report(
        "sandbox.user_metrics",
        "segment_name",
        db_key="analytics",
        metric_columns=["orders"],
        metric_names_override={"orders": "Orders per user"},
        groups_order=["control", "test_1", "test_all"],
        excel_file_name=output,
    )

    assert result["metric_name"].unique().tolist() == ["Orders per user"]
    workbook = load_workbook(output, read_only=True)
    assert workbook.sheetnames == ["summary", "raw_metrics"]
    summary_rows = list(workbook["summary"].iter_rows(values_only=True))
    assert summary_rows[1][:4] == ("metric", "control", "test_1", "test_all")
    assert summary_rows[3][0] == "Orders per user"
    raw_rows = list(workbook["raw_metrics"].iter_rows(values_only=True))
    assert raw_rows[0][0] == "segment_name"
    assert "Orders per user" in {row[4] for row in raw_rows[1:]}


def test_compute_metrics_report_writes_unsegmented_workbook(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    captured: dict[str, Any] = {}
    _install_sql_fakes(monkeypatch, captured=captured)
    output = tmp_path / "total.xlsx"

    result = reporting.compute_metrics_report(
        "sandbox.user_metrics",
        db_key="analytics",
        metric_columns=["orders"],
        excel_file_name=output,
    )

    assert "segment_name" not in result.columns
    workbook = load_workbook(output, read_only=True)
    assert workbook.sheetnames == ["summary", "raw_metrics"]
    assert next(iter(workbook["summary"].values))[:4] == (
        "metric",
        "control",
        "test_1",
        "test_all",
    )
    assert next(iter(workbook["raw_metrics"].values))[0] == "metric_type"


def test_compute_metrics_report_uses_pre_experiment_table(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    captured: dict[str, Any] = {}
    _install_sql_fakes(monkeypatch, captured=captured)

    reporting.compute_metrics_report(
        "sandbox.user_metrics",
        "segment_name",
        db_key="analytics",
        pre_exp_table_name="sandbox.user_metrics_pre",
        pre_exp_sql_where="period = 'pre'",
        metric_columns=["orders"],
        create_excel=False,
    )

    tasks = captured["tasks"]
    group_task = tasks["segment_0001_groups"]
    pooled_task = tasks["segment_0001_pooled"]
    assert group_task["pre_exp_source"] == "sandbox.user_metrics_pre"
    assert group_task["pre_exp_source_type"] == "table"
    assert "period = 'pre'" in str(group_task["pre_exp_sql_where"])
    assert pooled_task["pre_exp_source_type"] == "sql"
    assert "test_all" in str(pooled_task["pre_exp_source"])


@pytest.mark.parametrize(
    ("kwargs", "match"),
    [
        ({"groups_order": ["control", "control"]}, "must not contain duplicates"),
        ({"metric_names_override": {"missing": "Name"}}, "Unknown metric"),
        ({"all_segment_label": "north"}, "conflicts with an observed segment"),
        ({"pooled_test_group": "test_1"}, "conflicts with an observed group"),
    ],
)
def test_compute_metrics_report_validates_presentation_options(
    monkeypatch: pytest.MonkeyPatch,
    kwargs: dict[str, object],
    match: str,
) -> None:
    captured: dict[str, Any] = {}
    _install_sql_fakes(monkeypatch, captured=captured)

    with pytest.raises((TypeError, ValueError), match=match):
        reporting.compute_metrics_report(
            "sandbox.user_metrics",
            "segment_name",
            db_key="analytics",
            metric_columns=["orders"],
            create_excel=False,
            **kwargs,
        )


@pytest.mark.parametrize(
    ("updates", "error", "match"),
    [
        ({"segment": ""}, ValueError, "segment must be a non-empty string"),
        ({"group": ""}, ValueError, "group must be a non-empty string"),
        ({"group": "segment_name"}, ValueError, "must name different columns"),
        (
            {"segment": reporting._REPORT_SHEET_COLUMN},
            ValueError,
            "internal report column",
        ),
        ({"pooled_test_group": "control"}, ValueError, "must differ from control"),
        ({"create_excel": 1}, TypeError, "must be a boolean"),
        ({"report_significance_alpha": True}, TypeError, "must be a real number"),
        ({"report_significance_alpha": "0.1"}, TypeError, "must be a real number"),
        ({"report_significance_alpha": 0}, ValueError, "must be between 0 and 1"),
    ],
)
def test_validate_report_options_rejects_invalid_values(
    updates: dict[str, object],
    error: type[Exception],
    match: str,
) -> None:
    kwargs: dict[str, object] = {
        "segment": "segment_name",
        "group": "group_name",
        "control": "control",
        "user_id": "user_id",
        "pooled_test_group": "test_all",
        "all_segment_label": "ALL",
        "create_excel": False,
        "report_significance_alpha": 0.01,
    }
    kwargs.update(updates)

    with pytest.raises(error, match=match):
        reporting._validate_report_options(**kwargs)


@pytest.mark.parametrize(
    ("case", "match"),
    [
        ("missing_table", "does not exist"),
        ("missing_required", "Missing required column"),
        ("duplicate_metrics", "must not contain duplicates"),
        ("missing_metric", "Missing metric column"),
        ("no_metrics", "At least one metric"),
    ],
)
def test_resolve_report_source_validates_table_contract(
    monkeypatch: pytest.MonkeyPatch,
    case: str,
    match: str,
) -> None:
    info = _table_info("sandbox.metrics")
    metric_columns: list[str] | None = ["orders"]
    if case == "missing_table":
        info.exists = False
    elif case == "missing_required":
        del info.columns["user_id"]
    elif case == "duplicate_metrics":
        metric_columns = ["orders", "orders"]
    elif case == "missing_metric":
        metric_columns = ["missing"]
    else:
        info.columns = {
            "user_id": "bigint",
            "group_name": "text",
            "segment_name": "text",
        }
        metric_columns = None
    monkeypatch.setattr(reporting.sql, "table_info", lambda *_args: info)

    with pytest.raises(ValueError, match=match):
        reporting._resolve_report_source(
            db_key="analytics",
            table_name="sandbox.metrics",
            segment="segment_name",
            group="group_name",
            user_id="user_id",
            metric_columns=metric_columns,
            ratio_metrics=None,
        )


def test_resolve_report_source_uses_unresolved_table_fallback(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    info = _table_info("sandbox.metrics")
    info.resolved_table = None
    monkeypatch.setattr(reporting.sql, "table_info", lambda *_args: info)

    source = reporting._resolve_report_source(
        db_key="analytics",
        table_name="sandbox.metrics",
        segment="segment_name",
        group="group_name",
        user_id="user_id",
        metric_columns=["orders"],
        ratio_metrics=None,
    )

    assert source.table_sql == '"sandbox"."metrics"'


def test_ratio_component_columns_validates_and_deduplicates() -> None:
    with pytest.raises(ValueError, match="missing required key 'denominator'"):
        reporting._ratio_component_columns([{"numerator": "orders"}])

    assert reporting._ratio_component_columns(
        [{"numerator": "orders", "denominator": "orders"}]
    ) == ["orders"]


def test_read_distinct_values_requires_output_column(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    monkeypatch.setattr(reporting.sql, "read", lambda *_args, **_kwargs: pd.DataFrame())

    with pytest.raises(ValueError, match="did not return"):
        reporting._read_distinct_values(
            db_key="analytics",
            table_sql='"sandbox"."metrics"',
            column="segment_name",
            backend="gp",
            sql_where=None,
            output_column="segment_value",
            print_queries=False,
            retry_cnt=1,
            timeout_increment=0,
            query_label=None,
        )


def test_compute_metrics_report_rejects_mismatched_pre_backend(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    def fake_table_info(_db_key: str, table: str) -> SimpleNamespace:
        info = _table_info(table)
        if table.endswith("_pre"):
            info.backend = "trino"
        return info

    monkeypatch.setattr(reporting.sql, "table_info", fake_table_info)

    with pytest.raises(ValueError, match="same backend"):
        reporting.compute_metrics_report(
            "sandbox.metrics",
            "segment_name",
            db_key="analytics",
            pre_exp_table_name="sandbox.metrics_pre",
            metric_columns=["orders"],
            create_excel=False,
        )


def test_combine_report_results_handles_supported_and_invalid_results() -> None:
    frame = _metric_row("ALL", "test_1")
    combined = reporting._combine_report_results(frame)
    assert combined.equals(frame)
    assert combined is not frame

    with pytest.raises(RuntimeError, match="task_1: failed"):
        reporting._combine_report_results({"task_1": "failed"})
    with pytest.raises(ValueError, match="no dataframes"):
        reporting._combine_report_results({})


@pytest.mark.parametrize(
    ("overrides", "error", "match"),
    [
        (["orders"], TypeError, "must be a mapping"),
        ({1: "Orders"}, ValueError, "keys must be non-empty strings"),
        ({"orders": ""}, ValueError, "values must be non-empty strings"),
        (
            {"orders": "Combined", "revenue": "Combined"},
            ValueError,
            "create duplicate name",
        ),
    ],
)
def test_apply_metric_name_overrides_validates_mapping(
    overrides: object,
    error: type[Exception],
    match: str,
) -> None:
    frame = pd.DataFrame({"metric_name": ["orders", "revenue"]})

    with pytest.raises(error, match=match):
        reporting._apply_metric_name_overrides(frame, overrides)


@pytest.mark.parametrize("groups_order", ["control", ["control", ""]])
def test_resolve_group_order_validates_values(groups_order: object) -> None:
    with pytest.raises((TypeError, ValueError)):
        reporting._resolve_group_order(
            groups_order=groups_order,
            observed_groups=["control", "test_1"],
        )


def test_ordered_comparison_columns_deduplicates_simple_names() -> None:
    frame = pd.DataFrame(
        {
            "group_1": ["test_1", "test_1"],
            "group_2": ["control", "test_2"],
        }
    )

    assert reporting._ordered_comparison_columns(frame, simple_names=True) == ["test_1"]
    assert reporting._ordered_comparison_columns(frame, simple_names=False) == [
        "test_1_vs_control_delta_relative_significant",
        "test_1_vs_test_2_delta_relative_significant",
    ]


def test_resolve_excel_output_validates_and_builds_paths(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    monkeypatch.chdir(tmp_path)
    with pytest.raises(TypeError, match="string, Path, or None"):
        reporting._resolve_excel_output(
            table_name="sandbox.metrics",
            backend="gp",
            excel_file_name=1,
        )
    with pytest.raises(ValueError, match="must not be empty"):
        reporting._resolve_excel_output(
            table_name="sandbox.metrics",
            backend="gp",
            excel_file_name="",
        )

    assert (
        reporting._resolve_excel_output(
            table_name="sandbox.metrics",
            backend="gp",
            excel_file_name="report.xlsx",
        )
        == tmp_path / "report.xlsx"
    )
    assert (
        reporting._resolve_excel_output(
            table_name='sandbox."!!!"',
            backend="gp",
            excel_file_name=None,
        )
        == tmp_path / "ab_metrics.xlsx"
    )
